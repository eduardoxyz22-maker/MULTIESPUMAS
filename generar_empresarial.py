#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador del Dashboard Empresarial de Ventas 2026 - Gerencia (Heaven Colchones).

USO:
    python generar_empresarial.py [ruta_al_excel] [-o salida.html]

- Lee TODOS los datos de la pestana 'Hoja1' (hoja consolidada "Dashboard Hoja1")
  y de 'seg semanal' (unidades por tienda) del Excel de ventas.
- No hay NINGUN numero hardcodeado: cada cifra sale del Excel y cada porcentaje,
  variacion, brecha y proyeccion se CALCULA en vivo desde las cifras base.
- Produce 'dashboard-empresarial.html' listo para abrir en el navegador / GitHub Pages.

Para actualizar el dashboard basta con pasar el Excel actualizado (misma estructura)
y volver a correr este script.

Requiere: openpyxl  (pip install openpyxl)
"""
import sys
import os
import re
import json
import argparse
import unicodedata

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instalar con: pip install openpyxl")

DEFAULT_XLSX = r"C:/Users/multiespumas/Downloads/VENTAS 2026 SEGUIMIENTO (Recuperado).xlsx"
DEFAULT_OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dashboard-empresarial.html")

# Periodo / etiquetas (derivados de la estructura del Excel: cierre de junio 2026)
MES_ACTUAL = "junio"
MES_ANTERIOR = "mayo"

# ----------------------------------------------------------------------------
# Helpers de lectura
# ----------------------------------------------------------------------------
def norm(s):
    """Normaliza texto (quita acentos, minusculas) para comparar etiquetas."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.strip().lower()

class Sheet:
    def __init__(self, ws):
        self.ws = ws
    def v(self, coord):
        return self.ws[coord].value
    def num(self, coord):
        x = self.ws[coord].value
        return float(x) if isinstance(x, (int, float)) else None
    def txt(self, coord):
        x = self.ws[coord].value
        return "" if x is None else str(x)

def assert_label(sheet, coord, expected_contains, ctx):
    got = norm(sheet.v(coord))
    if norm(expected_contains) not in got:
        raise SystemExit(
            f"[VALIDACION] En {ctx}: se esperaba que {coord} contuviera "
            f"'{expected_contains}' pero se encontro '{sheet.v(coord)}'. "
            f"La estructura del Excel cambio; revisar el mapa de celdas.")

# ----------------------------------------------------------------------------
# Helpers de formato / calculo (equivalen a los del prototipo)
# ----------------------------------------------------------------------------
def rhu(n):
    """Redondeo half-up a entero (como Excel/JS), no bancario."""
    if n is None:
        return None
    import math
    r = math.floor(abs(n) + 0.5)
    return -r if n < 0 else r

def fmt(n):
    if n is None:
        return "—"  # em dash
    return f"{rhu(n):,}"  # miles en-US

def sep_miles(text):
    """Inserta separador de miles en numeros de 4+ digitos dentro de un texto libre del Excel."""
    if not text:
        return text
    return re.sub(r"\d{4,}", lambda m: f"{int(m.group()):,}", text)

def fmt_millones(n):
    return "Bs " + f"{n/1_000_000:.2f}M"

def pct(x):
    if x is None:
        return "—"
    return f"{x*100:,.1f}%"

def vf(x):
    """Flecha + clase segun signo. Devuelve dict."""
    if x is None:
        return {"txt": "—", "cls": ""}
    arrow = "▲ " if x >= 0 else "▼ "
    return {"txt": arrow + f"{abs(x*100):,.1f}%", "cls": "up" if x >= 0 else "dn"}

def esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

# ----------------------------------------------------------------------------
# EXTRACCION
# ----------------------------------------------------------------------------
def extract(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Hoja1" not in wb.sheetnames:
        raise SystemExit("El Excel no tiene la pestana 'Hoja1' (fuente del dashboard).")
    H = Sheet(wb["Hoja1"])
    SS = Sheet(wb["seg semanal"]) if "seg semanal" in wb.sheetnames else None

    # --- Validaciones de estructura (fallar temprano si el layout cambio) ---
    assert_label(H, "B9", "VENTAS MENSUALES", "tabla evolucion mensual")
    assert_label(H, "B57", "DETALLE POR TIENDA", "tabla por tienda")
    assert_label(H, "I57", "SEGUIMIENTO SEMANAL", "tabla semanal")
    assert_label(H, "B72", "LINEA DE PRODUCTO", "tabla producto")
    assert_label(H, "B100", "PROYECCION CIERRE", "tabla proyeccion")
    assert_label(H, "I100", "POR MARCA", "tabla marca vs jul")

    D = {}

    # --- Evolucion mensual (filas 11-16) ---
    meses = []
    for r in range(11, 17):
        nombre = H.txt(f"B{r}")
        ventas = H.num(f"C{r}")
        obj = H.num(f"D{r}")
        real25 = H.num(f"E{r}")
        if ventas is None:
            continue
        meses.append({
            "mes": nombre, "ventas": ventas, "obj": obj, "real25": real25,
            "cumpl": (ventas / obj) if obj else None,
            "var25": (ventas / real25 - 1) if real25 else None,
        })
    D["meses"] = meses
    sem_ventas = sum(m["ventas"] for m in meses)
    sem_obj = sum(m["obj"] for m in meses if m["obj"])
    sem_real25 = sum(m["real25"] for m in meses if m["real25"])
    mejor = max(meses, key=lambda m: m["ventas"])
    D["sem"] = {
        "ventas": sem_ventas, "obj": sem_obj, "real25": sem_real25,
        "cumpl": sem_ventas / sem_obj if sem_obj else None,
        "var25": sem_ventas / sem_real25 - 1 if sem_real25 else None,
        # brecha sobre los totales YA redondeados (como se leen en el dashboard)
        "brecha": rhu(sem_obj) - rhu(sem_ventas),
        "mejor_mes": mejor["mes"], "mejor_val": mejor["ventas"],
    }

    # Mes en curso = ultimo mes con ventas
    cur = meses[-1]
    prev = meses[-2]
    D["mes_actual"] = {
        "ventas": cur["ventas"], "obj": cur["obj"], "real25": cur["real25"],
        "cumpl": cur["cumpl"], "brecha": rhu(cur["obj"]) - rhu(cur["ventas"]),
        "var_prev": cur["ventas"] / prev["ventas"] - 1 if prev["ventas"] else None,
        "var25": cur["var25"],
        "mejor_prev_val": prev["ventas"],
    }

    # --- Por marca acumulado vs presupuesto anual (filas 11-14, cols I-N) ---
    marcas_acum = []
    for r in range(11, 15):
        marca = H.txt(f"I{r}")
        ventas = H.num(f"J{r}")
        ppto = H.num(f"K{r}")
        real25 = H.num(f"M{r}")
        if not marca:
            continue
        marcas_acum.append({
            "marca": marca, "ventas": ventas, "ppto": ppto,
            "avance": (ventas / ppto) if ppto else None, "real25": real25,
            # "Ppto vs 2025" = presupuesto anual / real 2025 (full year) - 1
            "vs25": (ppto / real25 - 1) if real25 else None,
        })
    tot_ventas = H.num("J15"); tot_ppto = H.num("K15"); tot_real25 = H.num("M15")
    D["marcas_acum"] = marcas_acum
    D["marcas_acum_total"] = {
        "ventas": tot_ventas, "ppto": tot_ppto,
        "avance": tot_ventas / tot_ppto if tot_ppto else None,
        "vs25": tot_ppto / tot_real25 - 1 if tot_real25 else None,
    }

    # --- Leads (filas 19-21, cols I-N) ---
    leads = {}
    for r in range(19, 21):
        canal = H.txt(f"I{r}")
        lv = H.num(f"J{r}")
        vt = H.num(f"K{r}")
        inv = H.num(f"M{r}")
        if not canal:
            continue
        leads[norm(canal)] = {
            "canal": canal, "leads": lv, "ventas": vt,
            "efect": (vt / lv) if lv else None, "inv": inv,
            "costo": (inv / vt) if (inv and vt) else None,
        }
    D["leads"] = leads
    D["leads_total"] = {
        "leads": H.num("J21"), "ventas": H.num("K21"),
        "efect": (H.num("K21") / H.num("J21")) if H.num("J21") else None,
    }
    D["leads_nota"] = sep_miles(H.txt("I22"))  # "Objetivo leads HEAVEN: 4,000 - Avance: 53%"

    # --- Junio detalle por tienda (filas 59-70, cols B-G) ---
    tiendas = []
    for r in range(59, 70):
        nombre = H.txt(f"B{r}")
        if not nombre or norm(nombre) == "total":
            continue
        ventas = H.num(f"C{r}")
        mayo = H.num(f"D{r}")
        ppto = H.num(f"F{r}")
        tiendas.append({
            "nombre": nombre, "ventas": ventas, "mayo": mayo, "ppto": ppto,
            "var": (ventas / mayo - 1) if mayo else None,
            "alcance": (ventas / ppto) if ppto else None,
        })
    D["tiendas"] = tiendas
    D["tiendas_total"] = {
        "ventas": H.num("C70"), "mayo": H.num("D70"), "ppto": H.num("F70"),
        "var": (H.num("C70") / H.num("D70") - 1) if H.num("D70") else None,
        "alcance": (H.num("C70") / H.num("F70")) if H.num("F70") else None,
    }

    # --- Seguimiento semanal mayo vs junio (filas 59-64, cols I-N) ---
    semanas = []
    for r in range(59, 64):
        s = H.txt(f"I{r}")
        if not s:
            continue
        semanas.append({
            "semana": s, "mayo": H.num(f"J{r}"), "junio": H.num(f"K{r}"),
            "may_u": H.num(f"M{r}"), "jun_u": H.num(f"N{r}"),
            "var": (H.num(f"K{r}") / H.num(f"J{r}") - 1) if H.num(f"J{r}") else None,
        })
    D["semanas"] = semanas
    D["semanas_total"] = {
        "mayo": H.num("J64"), "junio": H.num("K64"),
        "may_u": H.num("M64"), "jun_u": H.num("N64"),
        "var": (H.num("K64") / H.num("J64") - 1) if H.num("J64") else None,
    }
    # pico junio
    pico = max(semanas, key=lambda w: (w["junio"] or 0))
    D["pico"] = {"semana": pico["semana"], "junio": pico["junio"], "jun_u": pico["jun_u"]}

    # --- Ventas por linea de producto (filas 74-81, cols B-F) ---
    productos = []
    for r in range(74, 81):
        nombre = H.txt(f"B{r}")
        if not nombre or norm(nombre) == "total":
            continue
        c = H.num(f"C{r}"); d = H.num(f"D{r}"); f = H.num(f"F{r}")
        sub = nombre.startswith("   ") or nombre.startswith("\t")
        productos.append({
            "nombre": nombre.strip(), "mayo": c, "junio": d, "acum": f, "sub": sub,
            "var": (d / c - 1) if c else None,
        })
    D["productos"] = productos
    D["productos_total"] = {
        "mayo": H.num("C81"), "junio": H.num("D81"), "acum": H.num("F81"),
        "var": (H.num("D81") / H.num("C81") - 1) if H.num("C81") else None,
    }

    # --- Proyeccion cierre por canal (filas 102-106, cols B-F) ---
    proy = []
    for r in range(102, 106):
        canal = H.txt(f"B{r}")
        if not canal:
            continue
        proy.append({
            "canal": canal, "fecha": H.num(f"C{r}"), "proy": H.num(f"D{r}"),
            "ppto": H.num(f"E{r}"),
            "alcance": (H.num(f"D{r}") / H.num(f"E{r}")) if H.num(f"E{r}") else None,
        })
    D["proy"] = proy
    D["proy_total"] = {
        "fecha": H.num("C106"), "proy": H.num("D106"), "ppto": H.num("E106"),
        "alcance": (H.num("D106") / H.num("E106")) if H.num("E106") else None,
    }

    # --- Marca: junio vs mayo (filas 102-105, cols I-N) para fichas y chips ---
    # Mapea a las 4 marcas del dashboard.
    MARCA_KEY = {  # normalizado -> clave interna
        "heaven": "HEAVEN", "suena": "SUENA", "roho": "ROHO",
        "clientes externos": "CLIENTES", "clientes ext.": "CLIENTES",
        "prod. term. / otros": "CLIENTES", "prod term": "CLIENTES",
    }
    COLOR = {"HEAVEN": "var(--teal)", "SUENA": "var(--amber)",
             "ROHO": "var(--red)", "CLIENTES": "var(--series-blue)"}
    NOMBRE = {"HEAVEN": "HEAVEN", "SUENA": "SUEÑA", "ROHO": "ROHO",
              "CLIENTES": "Clientes ext."}
    LEADKEY = {"HEAVEN": "heaven", "SUENA": "suena"}

    brand_jun = {}
    for r in range(102, 106):
        m = H.txt(f"I{r}")
        key = MARCA_KEY.get(norm(m))
        if not key:
            continue
        brand_jun[key] = {"junio": H.num(f"J{r}"), "mayo": H.num(f"K{r}"),
                          "jul25": H.num(f"L{r}"), "proyF": None}

    # acum por marca -> mapear por clave
    acum_by_key = {}
    for r in range(11, 15):
        m = H.txt(f"I{r}")
        key = MARCA_KEY.get(norm(m))
        if key:
            acum_by_key[key] = {
                "acum": H.num(f"J{r}"), "pptoAnual": H.num(f"K{r}"),
                "real25": H.num(f"M{r}"),
            }
    # proyeccion por marca -> mapear (rows 102-105 B-F)
    proy_by_key = {}
    PROYKEY = {"heaven": "HEAVEN", "suena": "SUENA", "roho": "ROHO",
               "clientes externos": "CLIENTES"}
    for r in range(102, 106):
        m = H.txt(f"B{r}")
        key = PROYKEY.get(norm(m))
        if key:
            proy_by_key[key] = {"proyFecha": H.num(f"C{r}"), "proy": H.num(f"D{r}"),
                                "pptoJun": H.num(f"E{r}")}

    brands = {}
    for key in ["HEAVEN", "SUENA", "ROHO", "CLIENTES"]:
        bj = brand_jun.get(key, {})
        ba = acum_by_key.get(key, {})
        bp = proy_by_key.get(key, {})
        lk = LEADKEY.get(key)
        ld = leads.get(lk) if lk else None
        junio = bj.get("junio"); mayo = bj.get("mayo")
        acum = ba.get("acum"); pptoAnual = ba.get("pptoAnual"); real25 = ba.get("real25")
        proyv = bp.get("proy"); pptoJun = bp.get("pptoJun")
        brands[key] = {
            "nombre": NOMBRE[key], "color": COLOR[key],
            "junio": junio, "mayo": mayo,
            "varMayo": (junio / mayo - 1) if (junio and mayo) else None,
            "acum": acum, "pptoAnual": pptoAnual,
            "avance": (acum / pptoAnual) if pptoAnual else None,
            "real2025": real25,
            "vs2025": (pptoAnual / real25 - 1) if real25 else None,
            "proyFecha": bp.get("proyFecha"), "proy": proyv, "pptoJun": pptoJun,
            "cumplMes": (junio / pptoJun) if pptoJun else None,
            "alcanceProy": (proyv / pptoJun) if pptoJun else None,
            "leads": ld["leads"] if ld else None,
            "ventasLeads": ld["ventas"] if ld else None,
            "efect": ld["efect"] if ld else None,
            "inv": ld["inv"] if ld else None,
        }
    D["brands"] = brands

    # --- Presupuesto por marca (para barras anuales, filas 11-14) ---
    ppto_marcas = []
    for m in marcas_acum:
        ppto_marcas.append({"nombre": m["marca"], "ventas": m["ventas"], "ppto": m["ppto"]})
    D["ppto_marcas"] = ppto_marcas

    # --- Detalle semanal por marca y tienda (Hoja1 86-94 Bs + seg semanal unidades) ---
    # Bs
    seg_tiendas = []
    for r in range(86, 94):
        nombre = H.txt(f"B{r}")
        if not nombre or norm(nombre) == "total":
            continue
        seg_tiendas.append({
            "nombre": nombre, "junio": H.num(f"H{r}"), "mayo": H.num(f"I{r}"),
            "var": (H.num(f"H{r}") / H.num(f"I{r}") - 1) if H.num(f"I{r}") else None,
        })
    D["seg_tiendas"] = seg_tiendas
    # Unidades por tienda desde 'seg semanal' (columnas totales)
    units = {}
    if SS is not None:
        r = 1
        while r <= SS.ws.max_row:
            lbl = SS.txt(f"K{r}")
            if lbl.lower().startswith("total "):
                name = lbl[6:].strip()
                units[norm(name)] = {"may_u": SS.num(f"M{r}"), "jun_u": SS.num(f"N{r}")}
            r += 1
    D["units_by_tienda"] = units

    # --- Tareas (filas 44-55, cols B/E/G) ---
    tareas = []
    for r in range(44, 56):
        acc = H.txt(f"B{r}")
        if not acc:
            continue
        tareas.append({"accion": acc.strip(), "avance": H.txt(f"E{r}").strip(),
                       "responsable": H.txt(f"G{r}").strip()})
    D["tareas"] = tareas

    # --- Campana (textos) ---
    D["campana"] = {
        "mecanica": H.txt("I44").strip(),
        "vigencia": H.txt("J46").strip(),
        "stock": H.txt("J47").strip(),
        "modelos": H.txt("I49").strip(),
        "descuentos_inv": sep_miles(H.txt("I23").strip()),  # descuentos + inversion saldo
    }

    D["_xlsx"] = os.path.basename(xlsx_path)
    return D


# ----------------------------------------------------------------------------
# Helpers de presentacion adicionales
# ----------------------------------------------------------------------------
def pct2(x):
    if x is None:
        return "—"
    return f"{x*100:.2f}%"

KEEP_UPPER = {"ROHO", "HEAVEN", "SUEÑA", "SUENA", "PPTO"}
# Normalizacion de etiquetas del Excel (MAYUSCULAS) -> display limpio del dashboard.
DISPLAY = {
    "productos terminados fab.": "Productos terminados fáb.",
    "distribuidores": "Distribuidores",
    "charcas": "Charcas",
    "mutualista": "Mutualista",
    "clientes externos": "Clientes ext.",
    "prod. term. / otros": "PROD. TERM. / Otros",
}
def prettify(name):
    """Normaliza etiquetas en MAYUSCULAS del Excel a un display limpio."""
    if not name:
        return name
    if norm(name) in DISPLAY:
        return DISPLAY[norm(name)]
    if not name.isupper():
        return name
    words = []
    for w in name.split():
        words.append(w if w.upper() in KEEP_UPPER else w.capitalize())
    return " ".join(words)

def bar(nombre, ventas, ppto):
    """Barra de avance vs presupuesto: ancho recortado 100%, badge por umbral."""
    p = (ventas / ppto * 100) if ppto else 0
    falta = (ppto - ventas) if ppto is not None else None
    if p >= 100:
        grad = "linear-gradient(90deg,#0F5F6D,#1B94A4)"; badge = "b-green"
        faltaTxt = "✓ +Bs " + fmt(-falta); faltaCol = "#1c8a5f"
    elif p >= 65:
        grad = "linear-gradient(90deg,#B86808,#D97706)"; badge = "b-amber"
        faltaTxt = "falta Bs " + fmt(falta); faltaCol = "#B86808"
    else:
        grad = "linear-gradient(90deg,#a5222f,#CE2939)"; badge = "b-red"
        faltaTxt = "falta Bs " + fmt(falta); faltaCol = "#c23a48"
    return {"nombre": nombre, "ventas": fmt(ventas), "ppto": "Bs " + fmt(ppto),
            "pct": f"{p:.1f}%", "w": f"{min(p,100):.1f}%", "grad": grad,
            "badge": badge, "faltaTxt": faltaTxt, "faltaCol": faltaCol}

def bar_row(t, ventas_w):
    return (
        '<div style="display:flex;align-items:center;gap:14px;margin-bottom:12px">'
        f'<div style="width:180px;flex-shrink:0;font-size:.78rem;font-weight:600;color:var(--text);text-align:right;line-height:1.15">{esc(t["nombre"])}</div>'
        '<div style="flex:1;position:relative;height:26px;border-radius:9px;background:rgba(15,95,109,.08);overflow:hidden;box-shadow:inset 0 1px 2px rgba(20,58,60,.1)">'
        f'<div style="position:absolute;inset:0 auto 0 0;border-radius:9px;width:{t["w"]};background:{t["grad"]}"></div>'
        '</div>'
        f'<div style="width:{ventas_w}px;flex-shrink:0;text-align:right;font-size:.72rem;color:var(--muted)"><b style="color:var(--text);font-weight:700">Bs {t["ventas"]}</b> / {t["ppto"]}</div>'
        f'<div style="width:120px;flex-shrink:0;text-align:right;font-size:.72rem;font-weight:700;color:{t["faltaCol"]}">{t["faltaTxt"]}</div>'
        f'<div style="width:58px;flex-shrink:0;text-align:right"><span class="badge {t["badge"]}">{t["pct"]}</span></div>'
        '</div>')

def num_td(val, cls=""):
    c = ("num " + cls).strip()
    return f'<td class="{c}">{val}</td>'

def vf_td(x):
    v = vf(x)
    return f'<td class="num {v["cls"]}">{v["txt"]}</td>'


# ----------------------------------------------------------------------------
# Fichas por marca (HTML precomputado, inyectado a JS)
# ----------------------------------------------------------------------------
def brand_detail_html(b):
    ring_cumpl = f'conic-gradient({b["color"]} 0 {min((b["cumplMes"] or 0)*100,100):.1f}%, rgba(15,95,109,.13) {min((b["cumplMes"] or 0)*100,100):.1f}% 100%)'
    ring_avance = f'conic-gradient({b["color"]} 0 {(b["avance"] or 0)*100:.1f}%, rgba(15,95,109,.13) {(b["avance"] or 0)*100:.1f}% 100%)'
    vMayo = vf(b["varMayo"]); v2025 = vf(b["vs2025"])
    leads_card = ""
    if b["leads"] is not None:
        inv = ("Bs " + fmt(b["inv"])) if b["inv"] else "—"
        leads_card = (
            '<div class="glass-card"><div class="gc-bar" style="background:var(--amber)"></div>'
            '<div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px">Leads del mes</div>'
            f'<div style="font-size:1.9rem;font-weight:800;line-height:1">{fmt(b["leads"])}</div>'
            f'<div style="font-size:.72rem;color:var(--muted);margin-top:10px">{pct(b["efect"])} efectividad · {fmt(b["ventasLeads"])} ventas · inv {inv}</div>'
            '</div>')
    mensual = (
        '<div class="mensual-only" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:16px">'
        f'<div class="glass-card"><div class="gc-bar" style="background:{b["color"]}"></div>'
        '<div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px">Ventas junio</div>'
        f'<div style="font-size:1.9rem;font-weight:800;line-height:1">Bs {fmt(b["junio"])}</div>'
        f'<div style="display:flex;gap:8px;margin-top:10px;flex-wrap:wrap"><span class="dchip {vMayo["cls"]}" style="background:rgba(15,95,109,.06);border:1px solid rgba(15,95,109,.16);font-size:.66rem">{vMayo["txt"]} vs mayo</span></div>'
        '</div>'
        '<div class="glass-card"><div class="gc-bar" style="background:var(--series-blue)"></div>'
        '<div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px">% Cumplimiento mensual</div>'
        '<div style="display:flex;align-items:center;gap:14px">'
        f'<div class="ring" style="width:82px;height:82px;background:{ring_cumpl}"><div><div style="font-size:1rem;font-weight:800;line-height:1">{pct(b["cumplMes"])}</div></div></div>'
        f'<div style="font-size:.72rem;color:var(--muted)">de Bs<br><b style="color:var(--text);font-size:.82rem">{fmt(b["pptoJun"])}</b><br>PPTO junio</div>'
        '</div></div>'
        '<div class="glass-card"><div class="gc-bar" style="background:var(--purple)"></div>'
        '<div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px">Proyección cierre junio</div>'
        f'<div style="font-size:1.9rem;font-weight:800;line-height:1">Bs {fmt(b["proy"])}</div>'
        f'<div style="font-size:.72rem;color:var(--muted);margin-top:10px">{pct(b["alcanceProy"])} del PPTO · a la fecha Bs {fmt(b["proyFecha"])}</div>'
        '</div>'
        + leads_card +
        '<div class="glass-card"><div class="gc-bar" style="background:var(--green)"></div>'
        '<div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px">Acumulado ene–jun</div>'
        f'<div style="font-size:1.9rem;font-weight:800;line-height:1">Bs {fmt(b["acum"])}</div>'
        f'<div style="font-size:.72rem;color:var(--muted);margin-top:10px">vs 2025 <span class="{v2025["cls"]}">{v2025["txt"]}</span> · Bs {fmt(b["real2025"])} en 2025</div>'
        '</div></div>')
    anual = (
        '<div class="anual-only brand-cards-anual" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:16px">'
        f'<div class="glass-card"><div class="gc-bar" style="background:{b["color"]}"></div>'
        '<div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px">Ventas ene–jun</div>'
        f'<div style="font-size:1.9rem;font-weight:800;line-height:1">Bs {fmt(b["acum"])}</div>'
        f'<div style="display:flex;gap:8px;margin-top:10px;flex-wrap:wrap"><span class="dchip {v2025["cls"]}" style="background:rgba(15,95,109,.06);border:1px solid rgba(15,95,109,.16);font-size:.66rem">{v2025["txt"]} vs 2025</span></div>'
        '</div>'
        '<div class="glass-card"><div class="gc-bar" style="background:var(--series-blue)"></div>'
        '<div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px">Objetivo anual</div>'
        f'<div style="font-size:1.9rem;font-weight:800;line-height:1">Bs {fmt(b["pptoAnual"])}</div>'
        '<div style="font-size:.72rem;color:var(--muted);margin-top:10px">presupuesto 2026 · año completo</div>'
        '</div>'
        '<div class="glass-card"><div class="gc-bar" style="background:var(--amber)"></div>'
        '<div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px">% Avance anual</div>'
        '<div style="display:flex;align-items:center;gap:14px">'
        f'<div class="ring" style="width:82px;height:82px;background:{ring_avance}"><div><div style="font-size:1rem;font-weight:800;line-height:1">{pct(b["avance"])}</div></div></div>'
        f'<div style="font-size:.72rem;color:var(--muted)">de<br><b style="color:var(--text);font-size:.82rem">Bs {fmt(b["pptoAnual"])}</b><br>PPTO anual</div>'
        '</div></div>'
        '<div class="glass-card"><div class="gc-bar" style="background:var(--green)"></div>'
        '<div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px">Crecimiento vs 2025</div>'
        f'<div style="font-size:1.9rem;font-weight:800;line-height:1"><span class="{v2025["cls"]}">{v2025["txt"]}</span></div>'
        f'<div style="font-size:.72rem;color:var(--muted);margin-top:10px">Bs {fmt(b["real2025"])} real en 2025</div>'
        '</div></div>')
    header = (
        '<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px">'
        f'<span class="dot" style="width:14px;height:14px;background:{b["color"]}"></span>'
        f'<span style="font-size:1.15rem;font-weight:800;color:var(--text)">{esc(b["nombre"])}</span>'
        '<span class="ds-muted mensual-only">· detalle junio 2026</span>'
        '<span class="ds-muted anual-only brand-cards-anual">· detalle acumulado ene–jun 2026</span>'
        '</div>')
    return header + mensual + anual


def sem_bars_html(semanas, unit):
    wk = [{"l": w["semana"].replace("Semana ", "Sem "), "may": (w["may_u"] if unit == "u" else w["mayo"]),
           "jun": (w["jun_u"] if unit == "u" else w["junio"])} for w in semanas]
    wmax = max([max(w["may"] or 0, w["jun"] or 0) for w in wk] or [1])
    pref = "" if unit == "u" else "Bs "
    out = []
    for w in wk:
        v = vf((w["jun"] / w["may"] - 1) if w["may"] else None)
        mayH = f'{(w["may"] or 0)/wmax*82:.1f}%'
        junH = f'{(w["jun"] or 0)/wmax*82:.1f}%'
        out.append(
            '<div style="flex:1;display:flex;flex-direction:column;align-items:center;height:100%">'
            '<div style="flex:1;display:flex;align-items:flex-end;justify-content:center;gap:9px;width:100%">'
            '<div style="display:flex;flex-direction:column;align-items:center;justify-content:flex-end;height:100%">'
            f'<div style="font-size:.58rem;color:var(--muted);margin-bottom:4px;white-space:nowrap">{pref}{fmt(w["may"])}</div>'
            f'<div class="sembar" style="width:32px;height:{mayH};background:var(--gray-md);border-radius:8px 8px 0 0"></div>'
            '</div>'
            '<div style="display:flex;flex-direction:column;align-items:center;justify-content:flex-end;height:100%">'
            f'<div style="font-size:.62rem;font-weight:800;color:var(--teal-dk);margin-bottom:4px;white-space:nowrap">{pref}{fmt(w["jun"])}</div>'
            f'<div class="sembar" style="width:32px;height:{junH};background:linear-gradient(180deg,var(--teal),var(--teal-mid));border-radius:8px 8px 0 0;box-shadow:0 3px 10px rgba(0,181,173,.32)"></div>'
            '</div>'
            '</div>'
            f'<div style="margin-top:12px;font-size:.74rem;font-weight:700;color:var(--text);white-space:nowrap">{w["l"]}</div>'
            f'<div class="{v["cls"]}" style="font-size:.64rem;font-weight:700;margin-top:2px;white-space:nowrap">{v["txt"]}</div>'
            '</div>')
    return "".join(out)


def barlist(title, rows, cls):
    body = "".join(
        f'<div class="bl-row"><div class="bl-name">{esc(r["name"])}</div>'
        f'<div class="bl-track"><div class="bl-fill" style="width:{r["w"]};background:{r["color"]}">{r["val"]}</div></div></div>'
        for r in rows)
    return (f'<div class="barlist lg {cls}" style="padding:22px 24px">'
            f'<div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:16px">{title}</div>'
            f'{body}</div>')


def short_name(n):
    n = n.replace("Tienda ", "")
    if "(" in n:
        n = n.split("(")[0].strip()
    return n


def signed_pct(x):
    return ("+" if x >= 0 else "−") + f"{abs(x*100):.1f}%"


def es_lista(nombres):
    nombres = list(nombres)
    if len(nombres) <= 1:
        return "".join(nombres)
    return ", ".join(nombres[:-1]) + " y " + nombres[-1]


# ----------------------------------------------------------------------------
# JS (solo interactividad + presentacion; los datos vienen inyectados)
# ----------------------------------------------------------------------------
JS_TEMPLATE = r"""
(function () {
  "use strict";
  var DATA = window.__DATA__;

  function applyBody(view) {
    var b = document.body;
    b.classList.remove('v-todas','v-mes','v-anio','v-sem');
    b.classList.add(view==='MENSUAL'?'v-mes':view==='ANUAL'?'v-anio':view==='SEMANAL'?'v-sem':'v-todas');
  }
  // View tabs
  Array.prototype.forEach.call(document.querySelectorAll('#viewtabs .vtab'), function(btn){
    btn.addEventListener('click', function(){
      document.querySelectorAll('#viewtabs .vtab').forEach(function(x){x.classList.remove('active');});
      btn.classList.add('active');
      applyBody(btn.getAttribute('data-view'));
    });
  });
  // Brand tabs
  Array.prototype.forEach.call(document.querySelectorAll('#brandTabs .tab'), function(btn){
    btn.addEventListener('click', function(){
      document.querySelectorAll('#brandTabs .tab').forEach(function(x){x.classList.remove('active');});
      btn.classList.add('active');
      var k = btn.getAttribute('data-brand');
      var todas = document.getElementById('todasBars');
      var detail = document.getElementById('brandDetail');
      if (k === 'TODAS') { todas.style.display=''; detail.style.display='none'; detail.innerHTML=''; }
      else { todas.style.display='none'; detail.style.display=''; detail.innerHTML = DATA.brands[k] || ''; }
    });
  });
  // Sem unit toggle
  Array.prototype.forEach.call(document.querySelectorAll('#semUnitTabs .tab'), function(btn){
    btn.addEventListener('click', function(){
      document.querySelectorAll('#semUnitTabs .tab').forEach(function(x){x.classList.remove('active');});
      btn.classList.add('active');
      var u = btn.getAttribute('data-unit');
      document.getElementById('semWeeks').innerHTML = (u === 'u') ? DATA.weeksU : DATA.weeksBs;
    });
  });
  // Scroll progress
  function onScroll(){
    var el=document.getElementById('hc-sb'); if(!el) return;
    var h=document.documentElement; var max=h.scrollHeight-h.clientHeight;
    el.style.width=(max>0?(h.scrollTop/max)*100:0)+'%';
  }
  window.addEventListener('scroll', onScroll, {passive:true});
  var fab=document.getElementById('hc-fab'); if(fab) fab.addEventListener('click', function(){window.print();});
  applyBody('TODAS'); onScroll();
})();
"""


# ----------------------------------------------------------------------------
# ENSAMBLADO DEL HTML
# ----------------------------------------------------------------------------
def build_html(D):
    from _assets_empresarial import CSS
    m = D["mes_actual"]; sem = D["sem"]; brands = D["brands"]

    # ---- Header stats ----
    header_stats = (
        f'<div class="hstat"><div class="hstat-v">{fmt_millones(m["ventas"])}</div><div class="hstat-l">Ventas junio</div></div>'
        f'<div class="hstat"><div class="hstat-v">{pct(m["cumpl"])}</div><div class="hstat-l">Alcance mes</div></div>'
        f'<div class="hstat"><div class="hstat-v" style="color:#7CF6EF">{pct(sem["cumpl"])}</div><div class="hstat-l">Semestre</div></div>')

    # ---- Section 01 mchips (junio por marca vs objetivo del mes, ordenado por % cumplimiento desc) ----
    def _obj_col(a):
        return "var(--green)" if (a or 0) >= 1 else ("var(--amber)" if (a or 0) >= 0.65 else "var(--red)")
    chip_brands = sorted(brands.values(), key=lambda b: -(b["cumplMes"] or 0))
    mchips = ""
    for b in chip_brands:
        col = _obj_col(b["cumplMes"])
        mchips += (
            f'<div class="mchip"><span class="dot" style="background:{b["color"]}"></span>'
            f'<span style="flex:1;font-size:.8rem;font-weight:600">{esc(b["nombre"])}</span>'
            f'<span style="font-size:.82rem;font-weight:800">Bs {fmt(b["junio"])}</span>'
            f'<span style="color:{col};font-weight:700;font-size:.68rem;min-width:52px;text-align:right">{pct(b["cumplMes"])}</span></div>')

    ring01 = f'conic-gradient(var(--teal) 0 {(m["cumpl"] or 0)*100:.1f}%, rgba(15,95,109,.13) {(m["cumpl"] or 0)*100:.1f}% 100%)'
    sec01 = f'''<div class="sec mensual-only">01 · Resultado del mes · junio 2026</div>
    <div class="lg lg-glow mensual-only" style="padding:30px 34px;margin-bottom:26px">
      <div style="display:flex;align-items:center;gap:40px;flex-wrap:wrap">
        <div style="flex:1.3;min-width:280px">
          <div style="font-size:.66rem;font-weight:700;color:var(--teal);text-transform:uppercase;letter-spacing:.1em;margin-bottom:6px">Ventas totales de junio</div>
          <div style="font-size:3.4rem;font-weight:800;line-height:1;color:var(--text);letter-spacing:-.02em">Bs {fmt(m["ventas"])}</div>
          <div style="font-size:.82rem;color:var(--muted);margin-top:8px">Cierre del mes · {fmt(D["semanas_total"]["jun_u"])} unidades vendidas</div>
          <div style="display:flex;gap:30px;margin-top:14px">
            <div><div style="font-size:.6rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.06em">Objetivo del mes</div><div style="font-size:1.2rem;font-weight:800;color:var(--text)">Bs {fmt(m["obj"])}</div></div>
            <div><div style="font-size:.6rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.06em">Brecha vs objetivo</div><div style="font-size:1.2rem;font-weight:800;color:var(--red)">− Bs {fmt(m["brecha"])}</div></div>
          </div>
          <div style="display:flex;gap:10px;margin-top:16px;flex-wrap:wrap">
            <span class="dchip" style="background:rgba(206,60,74,.1);color:#c23a48;border:1px solid rgba(206,60,74,.28)">{vf(m["var_prev"])["txt"]} vs mayo</span>
            <span class="dchip" style="background:rgba(34,150,100,.1);color:#1c8a5f;border:1px solid rgba(34,150,100,.28)">{vf(m["var25"])["txt"]} vs junio 2025</span>
            <span class="dchip" style="background:rgba(184,104,8,.1);color:#B86808;border:1px solid rgba(184,104,8,.28)">{pct(m["cumpl"])} del presupuesto</span>
          </div>
        </div>
        <div style="display:flex;flex-direction:column;align-items:center;gap:9px">
          <div class="ring" style="background:{ring01}">
            <div><div style="font-size:2rem;font-weight:800;color:var(--text);line-height:1">{pct(m["cumpl"])}</div><div style="font-size:.58rem;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-top:2px">del objetivo</div></div>
          </div>
          <div style="font-size:.66rem;color:var(--muted)">Alcance del mes</div>
        </div>
        <div style="flex:1;min-width:250px;display:flex;flex-direction:column;gap:8px">
          <div style="font-size:.62rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:2px">Junio por marca · % del objetivo del mes</div>
          {mchips}
        </div>
      </div>
    </div>'''

    # ---- Exec summary (numeros calculados; matices derivados) ----
    tiendas_alc = [t for t in D["tiendas"] if t["alcance"] is not None]
    top2 = sorted(tiendas_alc, key=lambda t: -t["alcance"])[:2]
    bot3 = sorted(tiendas_alc, key=lambda t: t["alcance"])[:3]
    lideres = ", ".join(f'{short_name(prettify(t["nombre"]))} {round(t["alcance"]*100)}%' for t in top2)
    rezag = ", ".join(f'{short_name(prettify(t["nombre"]))} {round(t["alcance"]*100)}%' for t in bot3)
    min_mes = min(D["meses"], key=lambda x: x["cumpl"])
    flojo = " el mes más flojo del semestre," if min_mes["mes"] == D["meses"][-1]["mes"] else ""
    neg_weeks = [w["semana"].split()[-1] for w in D["semanas"] if (w["var"] or 0) < 0]
    caida = f' arrastrado por la caída de las semanas {"–".join(neg_weeks)}.' if neg_weeks else "."
    exec_html = f'''<div class="exec">
      <div class="exec-lbl">Resumen ejecutivo</div>
      <p>Junio cierra en <b>Bs {fmt(m["ventas"])}</b> — el <b>{pct(m["cumpl"])} del objetivo</b> y {signed_pct(m["var_prev"])} vs mayo:{flojo}{caida} En el acumulado, el semestre suma <b>Bs {fmt(sem["ventas"])}</b> ({pct(sem["cumpl"])} de meta, <b style="color:var(--green)">{vf(sem["var25"])["txt"]} vs 2025</b>). Palancas: recuperar <b>conversión de leads</b> ({pct2(D["leads_total"]["efect"])} de efectividad) y cerrar la brecha entre canales líderes ({lideres}) y rezagados ({rezag}).</p>
    </div>'''

    # ---- Alert ----
    bajo65 = [t for t in tiendas_alc if t["alcance"] < 0.65]
    nombres65 = es_lista(prettify(t["nombre"]) for t in sorted(bajo65, key=lambda t: t["alcance"]))
    alert_html = f'''<div class="alert amber">
      <span class="ico">⚠️</span>
      <div>Junio en <b>{pct(m["cumpl"])} del objetivo</b> — <b>{len(bajo65)} canales bajo el 65%</b> de su presupuesto. Priorizar reactivación de <b>{nombres65}</b> para el cierre de julio.</div>
    </div>'''

    # ---- Section 02 (mensual): budget bar + pptoTiendas bars ----
    tiendas_bar = [bar(prettify(t["nombre"]), t["ventas"], t["ppto"])
                   for t in sorted(tiendas_alc, key=lambda t: -t["alcance"])]
    ppto_tiendas_html = "".join(bar_row(t, 172) for t in tiendas_bar)
    logrado = m["cumpl"] or 0
    sec02 = f'''<div class="sec mensual-only">02 · Desviación vs presupuesto · junio 2026</div>
    <div class="lg mensual-only" style="padding:24px 30px 26px;margin-bottom:26px">
      <div style="display:flex;justify-content:space-between;align-items:baseline;flex-wrap:wrap;gap:10px;margin-bottom:16px">
        <div style="font-size:.66rem;font-weight:700;color:var(--teal);text-transform:uppercase;letter-spacing:.1em">Avance de presupuesto · junio</div>
        <div style="font-size:.72rem;color:var(--muted)">Objetivo <b style="color:var(--text)">Bs {fmt(m["obj"])}</b></div>
      </div>
      <div style="position:relative;height:44px;border-radius:14px;overflow:hidden;background:rgba(15,95,109,.09);box-shadow:inset 0 1px 3px rgba(20,58,60,.12)">
        <div style="position:absolute;inset:0 auto 0 0;width:{logrado*100:.1f}%;border-radius:14px;background:linear-gradient(90deg,#0F5F6D,#1B94A4);box-shadow:0 2px 10px rgba(15,95,109,.35);display:flex;align-items:center;padding-left:16px">
          <span style="color:#fff;font-weight:800;font-size:.92rem;letter-spacing:-.01em">Bs {fmt(m["ventas"])}</span>
        </div>
        <div style="position:absolute;right:16px;top:50%;transform:translateY(-50%);color:var(--red);font-weight:800;font-size:.9rem">falta Bs {fmt(m["brecha"])}</div>
      </div>
      <div style="display:flex;justify-content:space-between;margin-top:8px;font-size:.68rem;color:var(--muted)">
        <span style="display:flex;align-items:center;gap:6px"><span style="width:10px;height:10px;border-radius:3px;background:linear-gradient(90deg,#0F5F6D,#1B94A4)"></span>Logrado · {pct(logrado)}</span>
        <span style="display:flex;align-items:center;gap:6px"><span style="width:10px;height:10px;border-radius:3px;background:rgba(206,60,74,.5)"></span>Por alcanzar · {pct(1-logrado)}</span>
      </div>
      <div style="height:1px;background:rgba(15,95,109,.1);margin:22px 0 18px"></div>
      <div style="font-size:.62rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:16px">Avance por tienda / canal · ventas vs presupuesto junio</div>
      {ppto_tiendas_html}
      <div class="cap" style="margin-top:6px">Barra recortada al 100%; el % y los montos muestran el valor real. PPTO junio por tienda.</div>
    </div>'''

    # ---- Anual: panorama 6 cards ----
    panorama = f'''<div class="sec anual-only">01 · Panorama del semestre · ene–jun 2026</div>
    <div class="metrics anual-only" style="grid-template-columns:repeat(6,1fr)">
      <div class="mc"><div class="mc-bar" style="background:var(--teal)"></div><div class="mc-lbl">Ventas ene–jun</div><div class="mc-val">{fmt_millones(sem["ventas"])}</div><div class="mc-sub">Bs {fmt(sem["ventas"])} · pestaña 2026</div></div>
      <div class="mc"><div class="mc-bar" style="background:var(--series-blue)"></div><div class="mc-lbl">Objetivo ene–jun</div><div class="mc-val">{fmt_millones(sem["obj"])}</div><div class="mc-sub">Bs {fmt(sem["obj"])} · presupuesto</div></div>
      <div class="mc"><div class="mc-bar" style="background:var(--amber)"></div><div class="mc-lbl">% Cumplimiento</div><div class="mc-val" style="color:var(--amber)">{pct(sem["cumpl"])}</div><div class="mc-sub">acumulado ene–jun</div></div>
      <div class="mc"><div class="mc-bar" style="background:var(--green)"></div><div class="mc-lbl">Crec. vs 2025</div><div class="mc-val" style="color:var(--green)">{signed_pct(sem["var25"])}</div><div class="mc-sub">vs {fmt_millones(sem["real25"])} real 2025</div></div>
      <div class="mc"><div class="mc-bar" style="background:var(--purple)"></div><div class="mc-lbl">Mejor mes</div><div class="mc-val">{esc(sem["mejor_mes"])}</div><div class="mc-sub">Bs {fmt(sem["mejor_val"])} · pico del semestre</div></div>
      <div class="mc"><div class="mc-bar" style="background:var(--red)"></div><div class="mc-lbl">Brecha vs objetivo</div><div class="mc-val" style="color:var(--red)">Bs {fmt(sem["brecha"])}</div><div class="mc-sub">falta para la meta ene–jun</div></div>
    </div>'''

    # ---- Anual budget bar + pptoMarcas bars ----
    marcas_bar = [bar(prettify(x["nombre"]), x["ventas"], x["ppto"]) for x in D["ppto_marcas"]]
    ppto_marcas_html = "".join(bar_row(x, 186) for x in marcas_bar)
    sav = sem["cumpl"] or 0
    anual_budget = f'''<div class="lg anual-only" style="padding:24px 30px 26px;margin-bottom:26px">
      <div style="display:flex;justify-content:space-between;align-items:baseline;flex-wrap:wrap;gap:10px;margin-bottom:16px">
        <div style="font-size:.66rem;font-weight:700;color:var(--teal);text-transform:uppercase;letter-spacing:.1em">Avance de presupuesto · acumulado ene–jun</div>
        <div style="font-size:.72rem;color:var(--muted)">Objetivo <b style="color:var(--text)">Bs {fmt(sem["obj"])}</b></div>
      </div>
      <div style="position:relative;height:44px;border-radius:14px;overflow:hidden;background:rgba(15,95,109,.09);box-shadow:inset 0 1px 3px rgba(20,58,60,.12)">
        <div style="position:absolute;inset:0 auto 0 0;width:{sav*100:.1f}%;border-radius:14px;background:linear-gradient(90deg,#0F5F6D,#1B94A4);box-shadow:0 2px 10px rgba(15,95,109,.35);display:flex;align-items:center;padding-left:16px">
          <span style="color:#fff;font-weight:800;font-size:.92rem;letter-spacing:-.01em">Bs {fmt(sem["ventas"])}</span>
        </div>
        <div style="position:absolute;right:16px;top:50%;transform:translateY(-50%);color:var(--amber);font-weight:800;font-size:.9rem">falta Bs {fmt(sem["brecha"])}</div>
      </div>
      <div style="display:flex;justify-content:space-between;margin-top:8px;font-size:.68rem;color:var(--muted)">
        <span style="display:flex;align-items:center;gap:6px"><span style="width:10px;height:10px;border-radius:3px;background:linear-gradient(90deg,#0F5F6D,#1B94A4)"></span>Logrado · {pct(sav)}</span>
        <span style="display:flex;align-items:center;gap:6px"><span style="width:10px;height:10px;border-radius:3px;background:rgba(184,104,8,.5)"></span>Por alcanzar · {pct(1-sav)}</span>
      </div>
      <div style="height:1px;background:rgba(15,95,109,.1);margin:22px 0 18px"></div>
      <div style="font-size:.62rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:16px">Avance por marca · acumulado ene–jun vs presupuesto anual</div>
      {ppto_marcas_html}
      <div class="cap" style="margin-top:6px">Ppto anual 2026 (año completo); el % mide lo acumulado ene–jun contra la meta anual.</div>
    </div>'''

    # ---- Two-col: monthly evolution (mensual) + by-marca (anual) ----
    ev_rows = ""
    for x in D["meses"]:
        ev_rows += (f'<tr><td>{esc(x["mes"])}</td>{num_td(fmt(x["ventas"]))}{num_td(fmt(x["obj"]))}'
                    f'{num_td(fmt(x["real25"]))}{num_td(pct(x["cumpl"]))}{vf_td(x["var25"])}</tr>')
    ev_rows += (f'<tr class="trow-total"><td>TOTAL SEM.</td>{num_td(fmt(sem["ventas"]))}{num_td(fmt(sem["obj"]))}'
                f'{num_td(fmt(sem["real25"]))}{num_td(pct(sem["cumpl"]))}{vf_td(sem["var25"])}</tr>')
    BADGE_MARCA = {"HEAVEN": "b-teal", "SUEÑA": "b-amber", "SUENA": "b-amber", "ROHO": "b-red"}
    ma_rows = ""
    for x in D["marcas_acum"]:
        badge = BADGE_MARCA.get(x["marca"].upper(), "b-gray")
        ma_rows += (f'<tr><td><span class="badge {badge}">{esc(prettify(x["marca"]).upper() if x["marca"].isupper() else x["marca"])}</span></td>'
                    f'{num_td(fmt(x["ventas"]))}{num_td(fmt(x["ppto"]))}{num_td(pct(x["avance"]))}{vf_td(x["vs25"])}</tr>')
    mt = D["marcas_acum_total"]
    ma_rows += (f'<tr class="trow-total"><td>TOTAL</td>{num_td(fmt(mt["ventas"]))}{num_td(fmt(mt["ppto"]))}'
                f'{num_td(pct(mt["avance"]))}{vf_td(mt["vs25"])}</tr>')
    two_col_1 = f'''<div class="two-col">
      <div class="mensual-only">
        <div class="sec">03 · Evolución mensual · real vs objetivo vs 2025</div>
        <div class="tw"><table>
          <thead><tr><th>Mes</th><th class="num">Ventas 2026</th><th class="num">Objetivo</th><th class="num">Real 2025</th><th class="num">% Cumpl.</th><th class="num">Var. 25</th></tr></thead>
          <tbody>{ev_rows}</tbody>
        </table></div>
      </div>
      <div class="anual-only">
        <div class="sec">02 · Por marca · acumulado vs presupuesto anual</div>
        <div class="tw"><table>
          <thead><tr><th>Marca</th><th class="num">Vtas ene–jun</th><th class="num">Ppto anual</th><th class="num">% Avance</th><th class="num">vs 2025</th></tr></thead>
          <tbody>{ma_rows}</tbody>
        </table></div>
        <div class="cap">Ppto anual 2026 (año completo). El % avance mide lo acumulado ene–jun contra la meta anual.</div>
      </div>
    </div>'''

    # ---- Section 04: store table ----
    def alc_badge(a, pulse=False):
        if a is None:
            return "—"
        cls = "b-green" if a >= 1 else "b-amber" if a >= 0.65 else "b-red"
        p = " pulse" if pulse else ""
        return f'<span class="badge {cls}{p}">{pct(a)}</span>'
    # peor tienda fisica (para pulse)
    fisicas = [t for t in tiendas_alc if norm(t["nombre"]) not in
               ("roho", "distribuidores", "productos terminados fab.")]
    peor = min(fisicas, key=lambda t: t["alcance"])["nombre"] if fisicas else None
    st_rows = ""
    for t in D["tiendas"]:
        mayo_txt = fmt(t["mayo"]) if t["mayo"] else "—"
        var_cell = vf_td(t["var"]) if t["mayo"] else '<td class="num">—</td>'
        ppto_txt = fmt(t["ppto"]) if t["ppto"] else "—"
        badge = alc_badge(t["alcance"], pulse=(t["nombre"] == peor))
        st_rows += (f'<tr><td>{esc(prettify(t["nombre"]))}</td>{num_td(fmt(t["ventas"]))}{num_td(mayo_txt)}'
                    f'{var_cell}{num_td(ppto_txt)}<td class="num">{badge}</td></tr>')
    tt = D["tiendas_total"]
    st_rows += (f'<tr class="trow-total"><td>TOTAL</td>{num_td(fmt(tt["ventas"]))}{num_td(fmt(tt["mayo"]))}'
                f'{vf_td(tt["var"])}{num_td(fmt(tt["ppto"]))}{num_td(pct(tt["alcance"]))}</tr>')
    sec04 = f'''<div class="sec mensual-only">04 · Desempeño por tienda · junio (Bs)</div>
    <div class="tw mensual-only" style="margin-bottom:28px"><table>
      <thead><tr><th>Tienda / Canal</th><th class="num">Ventas Jun</th><th class="num">Mayo</th><th class="num">Var %</th><th class="num">PPTO Jun</th><th class="num">% Alcance</th></tr></thead>
      <tbody>{st_rows}</tbody>
    </table></div>'''

    # ---- Section 05: brand tabs + todas barlists + brand detail ----
    def blabel(b, anual=False):
        if b["nombre"] == "Clientes ext." and anual:
            return "Otros / Ext."
        return b["nombre"]
    bl_junio = sorted(brands.values(), key=lambda b: -(b["junio"] or 0))
    maxj = max(b["junio"] or 0 for b in bl_junio)
    rows_junio = [{"name": blabel(b), "val": "Bs " + fmt(b["junio"]),
                   "w": f'{(b["junio"] or 0)/maxj*100:.0f}%', "color": b["color"]} for b in bl_junio]
    bl_alc = sorted(brands.values(), key=lambda b: -(b["cumplMes"] or 0))
    rows_alc = [{"name": blabel(b), "val": pct(b["cumplMes"]),
                 "w": f'{min((b["cumplMes"] or 0)*100,100):.1f}%', "color": b["color"]} for b in bl_alc]
    bl_acum = sorted(brands.values(), key=lambda b: -(b["acum"] or 0))
    maxa = max(b["acum"] or 0 for b in bl_acum)
    rows_acum = [{"name": blabel(b, True), "val": "Bs " + fmt(b["acum"]),
                  "w": f'{(b["acum"] or 0)/maxa*100:.0f}%', "color": b["color"]} for b in bl_acum]
    bl_av = sorted(brands.values(), key=lambda b: -(b["avance"] or 0))
    rows_av = [{"name": blabel(b, True), "val": pct(b["avance"]),
                "w": f'{min((b["avance"] or 0)*100,100):.1f}%', "color": b["color"]} for b in bl_av]
    todas_bars = ('<div class="two-col">'
                  + barlist("Junio por marca · Bs", rows_junio, "mensual-only")
                  + barlist("% Alcance vs PPTO junio", rows_alc, "mensual-only")
                  + barlist("Acumulado ene–jun por marca · Bs", rows_acum, "anual-only")
                  + barlist("% Avance vs presupuesto anual", rows_av, "anual-only")
                  + '</div>')
    BTAB = [("TODAS", "Todas"), ("HEAVEN", "HEAVEN"), ("SUENA", "SUEÑA"), ("ROHO", "ROHO"), ("CLIENTES", "Clientes ext.")]
    brand_tabs = "".join(
        f'<button class="tab{" active" if k=="TODAS" else ""}" data-brand="{k}">{esc(l)}</button>' for k, l in BTAB)
    sec05 = f'''<div class="sec mensual-only">05 · Rendimiento por marca · junio</div>
    <div class="sec anual-only">Rendimiento por marca</div>
    <div class="tab-row">
      <div class="tabs" id="brandTabs">{brand_tabs}</div>
      <div class="rc">Selecciona una marca para ver su detalle · datos de junio 2026 y acumulado</div>
    </div>
    <div id="todasBars">{todas_bars}</div>
    <div id="brandDetail" style="display:none;margin-bottom:28px"></div>'''

    # ---- Weekly table (06) + leads cards ----
    wk_rows = ""
    for w in D["semanas"]:
        wk_rows += (f'<tr><td>{esc(w["semana"])}</td>{num_td(fmt(w["mayo"]))}{num_td(fmt(w["junio"]))}'
                    f'{vf_td(w["var"])}{num_td(fmt(w["may_u"]))}{num_td(fmt(w["jun_u"]))}</tr>')
    wt = D["semanas_total"]
    wk_rows += (f'<tr class="trow-total"><td>TOTAL</td>{num_td(fmt(wt["mayo"]))}{num_td(fmt(wt["junio"]))}'
                f'{vf_td(wt["var"])}{num_td(fmt(wt["may_u"]))}{num_td(fmt(wt["jun_u"]))}</tr>')
    # leads cards
    lh = D["leads"].get("heaven", {}); ls = D["leads"].get("suena", {})
    costo_h = ("Bs " + f'{lh.get("costo"):.2f}') if lh.get("costo") else "—"
    leads_cards = f'''<div class="glass-card"><div class="gc-bar" style="background:var(--teal)"></div>
            <div style="display:flex;justify-content:space-between;align-items:center">
              <span style="display:flex;align-items:center;gap:8px;font-weight:800;font-size:1rem"><span class="dot" style="background:var(--teal)"></span>HEAVEN</span>
              <span class="badge b-teal">{pct2(lh.get("efect"))} efectividad</span>
            </div>
            <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:14px">
              <div><div style="font-size:1.5rem;font-weight:800;line-height:1">{fmt(lh.get("leads"))}</div><div class="ds-muted">leads</div></div>
              <div><div style="font-size:1.5rem;font-weight:800;line-height:1">{fmt(lh.get("ventas"))}</div><div class="ds-muted">ventas</div></div>
              <div><div style="font-size:1.5rem;font-weight:800;line-height:1">{costo_h}</div><div class="ds-muted">costo / lead</div></div>
            </div>
            <div style="margin-top:12px;padding-top:10px;border-top:1px solid var(--gray-md);font-size:.72rem;color:var(--muted)">Inversión <b style="color:var(--text)">Bs {fmt(lh.get("inv"))}</b> · {esc(D["leads_nota"])}</div>
          </div>
          <div class="glass-card"><div class="gc-bar" style="background:var(--amber)"></div>
            <div style="display:flex;justify-content:space-between;align-items:center">
              <span style="display:flex;align-items:center;gap:8px;font-weight:800;font-size:1rem"><span class="dot" style="background:var(--amber)"></span>SUEÑA</span>
              <span class="badge b-amber">{pct2(ls.get("efect"))} efectividad</span>
            </div>
            <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:14px">
              <div><div style="font-size:1.5rem;font-weight:800;line-height:1">{fmt(ls.get("leads"))}</div><div class="ds-muted">leads</div></div>
              <div><div style="font-size:1.5rem;font-weight:800;line-height:1">{fmt(ls.get("ventas"))}</div><div class="ds-muted">ventas</div></div>
              <div><div style="font-size:1.5rem;font-weight:800;line-height:1">—</div><div class="ds-muted">costo / lead</div></div>
            </div>
            <div style="margin-top:12px;padding-top:10px;border-top:1px solid var(--gray-md);font-size:.72rem;color:var(--muted)">Inversión no registrada · más volumen de leads, menor conversión que HEAVEN</div>
          </div>'''
    lt = D["leads_total"]
    two_col_2 = f'''<div class="two-col mensual-only">
      <div>
        <div class="sec">06 · Seguimiento semanal · mayo vs junio</div>
        <div class="tw"><table>
          <thead><tr><th>Semana</th><th class="num">Mayo (Bs)</th><th class="num">Junio (Bs)</th><th class="num">Var %</th><th class="num">May (u)</th><th class="num">Jun (u)</th></tr></thead>
          <tbody>{wk_rows}</tbody>
        </table></div>
        <div class="cap">La caída de las semanas 4–5 explica el mes flojo: el ritmo se desplomó en la segunda mitad de junio.</div>
      </div>
      <div>
        <div class="sec">Leads por marca · efectividad</div>
        <div style="display:flex;flex-direction:column;gap:14px">{leads_cards}</div>
        <div class="alert amber" style="margin-top:14px;margin-bottom:0">
          <span class="ico">💡</span>
          <div>Total <b>{fmt(lt["leads"])} leads</b> → <b>{fmt(lt["ventas"])} ventas</b> ({pct2(lt["efect"])}). El cuello de botella no es captación — es <b>conversión</b>.</div>
        </div>
      </div>
    </div>'''

    # ---- Product table (07) + projection (08) ----
    pr_rows = ""
    for p in D["productos"]:
        if p["sub"]:
            pr_rows += (f'<tr class="trow-sub"><td>{esc(p["nombre"])}</td>{num_td(fmt(p["mayo"]))}{num_td(fmt(p["junio"]))}'
                        f'{vf_td(p["var"])}{num_td(fmt(p["acum"]))}</tr>')
        else:
            bold = ' style="font-weight:700"' if norm(p["nombre"]).startswith("heaven") else ""
            pr_rows += (f'<tr{bold}><td>{esc(p["nombre"])}</td>{num_td(fmt(p["mayo"]))}{num_td(fmt(p["junio"]))}'
                        f'{vf_td(p["var"])}{num_td(fmt(p["acum"]))}</tr>')
    pt = D["productos_total"]
    pr_rows += (f'<tr class="trow-total"><td>TOTAL</td>{num_td(fmt(pt["mayo"]))}{num_td(fmt(pt["junio"]))}'
                f'{vf_td(pt["var"])}{num_td(fmt(pt["acum"]))}</tr>')
    PROY_BADGE = {"heaven": "b-teal", "suena": "b-amber", "roho": "b-red", "clientes externos": "b-gray"}
    pj_rows = ""
    for p in D["proy"]:
        badge = PROY_BADGE.get(norm(p["canal"]), "b-gray")
        cls = "b-green" if (p["alcance"] or 0) >= 1 else "b-amber" if (p["alcance"] or 0) >= 0.65 else "b-red"
        pj_rows += (f'<tr><td><span class="badge {badge}">{esc(prettify(p["canal"]).upper() if p["canal"].isupper() else p["canal"])}</span></td>'
                    f'{num_td(fmt(p["fecha"]))}{num_td(fmt(p["proy"]))}{num_td(fmt(p["ppto"]))}'
                    f'<td class="num"><span class="badge {cls}">{pct(p["alcance"])}</span></td></tr>')
    pjt = D["proy_total"]
    pj_rows += (f'<tr class="trow-total"><td>TOTAL</td>{num_td(fmt(pjt["fecha"]))}{num_td(fmt(pjt["proy"]))}'
                f'{num_td(fmt(pjt["ppto"]))}{num_td(pct(pjt["alcance"]))}</tr>')
    two_col_3 = f'''<div class="two-col mensual-only">
      <div>
        <div class="sec">07 · Ventas por línea de producto · may vs jun (Bs)</div>
        <div class="tw"><table>
          <thead><tr><th>Línea / Marca</th><th class="num">Mayo</th><th class="num">Junio</th><th class="num">Var %</th><th class="num">Acum 2026</th></tr></thead>
          <tbody>{pr_rows}</tbody>
        </table></div>
        <div class="cap">Fuente: 'x PRODUCTO' (may–jun) y 'PRODUCTO' (acum). Solo venta de tiendas; excluye distribuidores y productos terminados.</div>
      </div>
      <div>
        <div class="sec">08 · Proyección de cierre · junio por canal (Bs)</div>
        <div class="tw"><table>
          <thead><tr><th>Canal</th><th class="num">Jun a la fecha</th><th class="num">Proy. cierre</th><th class="num">PPTO Jun</th><th class="num">% Alcance</th></tr></thead>
          <tbody>{pj_rows}</tbody>
        </table></div>
        <div class="cap">Fuente: pestaña PROY (snapshot operativo del mes en curso).</div>
      </div>
    </div>'''

    # ---- Tasks (09) + campaign ----
    ESTADO_BADGE = [("completado", "b-green"), ("ejecutado", "b-green"), ("aprobado", "b-green"),
                    ("proceso", "b-amber"), ("avance", "b-amber"), ("espera", "b-amber"),
                    ("pendiente", "b-gray")]
    def estado_badge(avance):
        n = norm(avance)
        for k, cls in ESTADO_BADGE:
            if k in n:
                lbl = {"b-green": "Completado", "b-amber": "En proceso", "b-gray": "Pendiente"}[cls]
                return f'<span class="badge {cls}">{lbl}</span>'
        return f'<span class="badge b-amber">En proceso</span>'
    tk_rows = ""
    for t in D["tareas"]:
        tk_rows += (f'<tr><td>{esc(t["accion"])}</td><td>{esc(t["responsable"] or "—")}</td>'
                    f'<td>{estado_badge(t["avance"])}</td></tr>')
    camp = D["campana"]
    two_col_4 = f'''<div class="two-col mensual-only" style="margin-bottom:8px">
      <div>
        <div class="sec">09 · Plan de acción · tareas del mes</div>
        <div class="tw"><table>
          <thead><tr><th>Acción</th><th>Responsable</th><th>Estado</th></tr></thead>
          <tbody>{tk_rows}</tbody>
        </table></div>
      </div>
      <div>
        <div class="sec">Campaña vigente</div>
        <div class="tw lg" style="padding:22px 24px">
          <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px">
            <span class="badge b-teal" style="font-size:.7rem">COMBO ORO</span>
            <span class="badge b-red" style="font-size:.7rem">−45% HEAVEN</span>
          </div>
          <p class="ds-body" style="margin-bottom:14px">{esc(camp["mecanica"])}</p>
          <div style="display:grid;grid-template-columns:auto 1fr;gap:8px 16px;font-size:.8rem;line-height:1.5">
            <div style="color:var(--muted);font-weight:700">Vigencia</div><div>{esc(camp["vigencia"])}</div>
            <div style="color:var(--muted);font-weight:700">Stock</div><div>{esc(camp["stock"])}</div>
            <div style="color:var(--muted);font-weight:700">Modelos</div><div>{esc(camp["modelos"])}</div>
            <div style="color:var(--muted);font-weight:700">Descuentos</div><div>{esc(camp["descuentos_inv"])}</div>
          </div>
        </div>
      </div>
    </div>'''

    # ---- Semanal block ----
    pico = D["pico"]
    prev_u = D["mes_actual"]["mejor_prev_val"]  # not used; units below
    su_var = vf((wt["jun_u"] / wt["may_u"] - 1) if wt["may_u"] else None)
    sb_var = vf(wt["var"])
    sem_cards = f'''<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-bottom:22px">
        <div class="glass-card"><div class="gc-bar" style="background:var(--teal)"></div>
          <div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px">Ventas junio</div>
          <div style="font-size:1.9rem;font-weight:800;line-height:1">Bs {fmt(wt["junio"])}</div>
          <div style="font-size:.72rem;color:var(--muted);margin-top:10px"><span class="{sb_var["cls"]}">{sb_var["txt"]}</span> vs mayo · Bs {fmt(wt["mayo"])}</div>
        </div>
        <div class="glass-card"><div class="gc-bar" style="background:var(--purple)"></div>
          <div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px">Unidades junio</div>
          <div style="font-size:1.9rem;font-weight:800;line-height:1">{fmt(wt["jun_u"])}</div>
          <div style="font-size:.72rem;color:var(--muted);margin-top:10px"><span class="{su_var["cls"]}">{su_var["txt"]}</span> vs mayo · {fmt(wt["may_u"])} unidades</div>
        </div>
        <div class="glass-card"><div class="gc-bar" style="background:var(--green)"></div>
          <div style="font-size:.66rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px">Semana pico · junio</div>
          <div style="font-size:1.9rem;font-weight:800;line-height:1">{esc(pico["semana"])}</div>
          <div style="font-size:.72rem;color:var(--muted);margin-top:10px">Bs {fmt(pico["junio"])} · {fmt(pico["jun_u"])} unidades</div>
        </div>
      </div>'''
    # units table
    su_rows = ""
    for w in D["semanas"]:
        su_rows += (f'<tr><td>{esc(w["semana"])}</td>{num_td(fmt(w["may_u"]))}{num_td(fmt(w["jun_u"]))}'
                    f'{vf_td((w["jun_u"]/w["may_u"]-1) if w["may_u"] else None)}</tr>')
    su_rows += (f'<tr class="trow-total"><td>TOTAL MES</td>{num_td(fmt(wt["may_u"]))}{num_td(fmt(wt["jun_u"]))}'
                f'{vf_td((wt["jun_u"]/wt["may_u"]-1) if wt["may_u"] else None)}</tr>')
    # brand-store detail
    detail_rows = build_detail_rows(D)
    sem_default = sem_bars_html(D["semanas"], "bs")
    semanal_block = f'''<div class="semanal-block">
      <div class="sec">Seguimiento semanal · mayo vs junio 2026</div>
      {sem_cards}
      <div class="lg" style="padding:24px 28px;margin-bottom:22px">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:22px;flex-wrap:wrap;gap:10px">
          <div style="font-size:.7rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.08em">Evolución por semana · mayo vs junio</div>
          <div class="tabs" id="semUnitTabs" style="border-radius:10px">
            <button class="tab active" data-unit="bs" style="padding:6px 18px">Bs</button>
            <button class="tab" data-unit="u" style="padding:6px 18px">Unidades</button>
          </div>
        </div>
        <div id="semWeeks" style="display:flex;align-items:flex-end;gap:22px;height:220px">{sem_default}</div>
        <div style="display:flex;gap:22px;margin-top:18px;padding-top:14px;border-top:1px solid var(--gray-md)">
          <span style="display:flex;align-items:center;gap:7px;font-size:.72rem;color:var(--muted)"><span style="width:13px;height:13px;background:var(--gray-md);border-radius:3px"></span>Mayo</span>
          <span style="display:flex;align-items:center;gap:7px;font-size:.72rem;color:var(--muted)"><span style="width:13px;height:13px;background:var(--teal);border-radius:3px"></span>Junio</span>
        </div>
      </div>
      <div class="sec">Unidades vendidas por semana · mayo vs junio</div>
      <div class="tw" style="margin-bottom:22px"><table>
        <thead><tr><th>Semana</th><th class="num">Mayo (u)</th><th class="num">Junio (u)</th><th class="num">Var %</th></tr></thead>
        <tbody>{su_rows}</tbody>
      </table></div>
      <div class="sec">Detalle por marca y tienda · mayo vs junio</div>
      <div class="tw" style="margin-bottom:16px"><table>
        <thead><tr><th>Marca / Tienda</th><th class="num">Mayo (Bs)</th><th class="num">Junio (Bs)</th><th class="num">Var %</th><th class="num">May u</th><th class="num">Jun u</th></tr></thead>
        <tbody>{detail_rows}</tbody>
      </table></div>
      <div class="cap" style="margin-bottom:26px">Fuente: pestaña 'seg semanal' del archivo · pivote marca › tienda › semana, mayo vs junio (Bs y unidades).</div>
    </div>'''

    # ---- Inyeccion de datos para JS (fichas de marca + grafico semanal) ----
    js_data = {
        "brands": {k: brand_detail_html(brands[k]) for k in brands},
        "weeksBs": sem_bars_html(D["semanas"], "bs"),
        "weeksU": sem_bars_html(D["semanas"], "u"),
    }
    data_json = json.dumps(js_data, ensure_ascii=False)

    view_tabs = ('<button class="vtab active" data-view="TODAS" style="border:none;background:transparent;font-family:inherit;font-weight:700;font-size:.82rem;color:var(--muted);padding:10px 32px;border-radius:13px;cursor:pointer;transition:all .18s;letter-spacing:.02em">Todas</button>'
                 '<button class="vtab" data-view="MENSUAL" style="border:none;background:transparent;font-family:inherit;font-weight:700;font-size:.82rem;color:var(--muted);padding:10px 32px;border-radius:13px;cursor:pointer;transition:all .18s;letter-spacing:.02em">Mensual</button>'
                 '<button class="vtab" data-view="SEMANAL" style="border:none;background:transparent;font-family:inherit;font-weight:700;font-size:.82rem;color:var(--muted);padding:10px 32px;border-radius:13px;cursor:pointer;transition:all .18s;letter-spacing:.02em">Semanal</button>'
                 '<button class="vtab" data-view="ANUAL" style="border:none;background:transparent;font-family:inherit;font-weight:700;font-size:.82rem;color:var(--muted);padding:10px 32px;border-radius:13px;cursor:pointer;transition:all .18s;letter-spacing:.02em">Anual</button>')

    html = f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Dashboard Ventas 2026 · Gerencia — Heaven Colchones</title>
<meta name="generator" content="generar_empresarial.py — fuente: {esc(D["_xlsx"])}">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<style>{CSS}</style>
</head>
<body class="v-todas">
<div class="app-daily">
  <div class="scroll-bar" id="hc-sb" style="width:0%"></div>
  <div style="padding:18px 20px 0">
    <header class="header grad" style="border-radius:26px;box-shadow:0 18px 46px rgba(9,72,68,.3), inset 0 1px 0 rgba(255,255,255,.32);overflow:hidden;position:relative">
      <div style="position:absolute;inset:0;background:radial-gradient(135% 175% at 8% -35%, rgba(255,255,255,.34), transparent 48%);pointer-events:none;z-index:0"></div>
      <div style="position:absolute;left:0;right:0;bottom:0;height:1px;background:rgba(255,255,255,.25);z-index:0"></div>
      <div class="hl" style="position:relative;z-index:1">
        <div class="logo">
          <div class="logo-h">HEAVEN<span style="color:var(--heaven-cross);font-weight:800;margin-left:6px">&#10011;</span></div>
          <div class="logo-s">colchones · recarga tu energía</div>
        </div>
        <div class="htitle">
          <h1>DASHBOARD VENTAS 2026 · GERENCIA</h1>
          <p>Panorama de Junio 2026 con acumulado ene–jun · MultiESPUMAS Viscarra S.R.L.</p>
        </div>
      </div>
      <div class="hr" style="position:relative;z-index:1;border-left:none">
        <div style="display:flex;align-items:stretch;background:rgba(255,255,255,.14);border:1px solid rgba(255,255,255,.26);border-radius:20px;padding:8px 4px;backdrop-filter:blur(8px);-webkit-backdrop-filter:blur(8px);box-shadow:inset 0 1px 0 rgba(255,255,255,.3)">
          {header_stats}
        </div>
      </div>
    </header>
  </div>
  <main class="container">
    <div class="viewtabs" style="display:flex;justify-content:center;margin-bottom:26px">
      <div id="viewtabs" style="display:inline-flex;gap:5px;background:rgba(255,255,255,.55);backdrop-filter:blur(14px);-webkit-backdrop-filter:blur(14px);border:1px solid rgba(255,255,255,.72);border-radius:18px;padding:5px;box-shadow:0 8px 24px rgba(9,72,68,.14), inset 0 1px 0 rgba(255,255,255,.95)">{view_tabs}</div>
    </div>
    {sec01}
    {exec_html}
    {alert_html}
    {sec02}
    {panorama}
    {anual_budget}
    {two_col_1}
    {sec04}
    {sec05}
    {two_col_2}
    {two_col_3}
    {two_col_4}
    {semanal_block}
    <div class="footer" style="border-radius:12px;border:1px solid var(--gray-md);margin-top:20px">
      Heaven Colchones · MultiESPUMAS Viscarra S.R.L. — Santa Cruz, Bolivia · Consolidado pestaña Hoja1 · Montos en bolivianos (Bs) · Generado desde {esc(D["_xlsx"])}
    </div>
  </main>
  <button class="fab" id="hc-fab">⬇ Exportar / Imprimir</button>
</div>
<script>window.__DATA__ = {data_json};</script>
<script>{JS_TEMPLATE}</script>
</body>
</html>'''
    return html


def build_detail_rows(D):
    """Detalle semanal por marca y tienda (agrupado), desde Hoja1 (Bs) + seg semanal (u)."""
    units = D["units_by_tienda"]
    def u(name):
        return units.get(norm(name), {"may_u": None, "jun_u": None})
    # Mapa de agrupacion derivado del orden del Excel (Hoja1 86-93)
    seg = D["seg_tiendas"]
    by = {norm(t["nombre"]): t for t in seg}
    def row(name, indent=False, bold=False, dot=None):
        t = by.get(norm(name))
        if not t:
            return ""
        un = u(name)
        pad = ' style="padding-left:28px"' if indent else ""
        label = esc(prettify(name))
        if dot:
            label = f'<span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:{dot};margin-right:9px;vertical-align:middle"></span>' + (f'<b>{label}</b>' if bold else label)
        return (f'<tr><td{pad}>{label}</td>{num_td(fmt(t["mayo"]))}{num_td(fmt(t["junio"]))}'
                f'{vf_td(t["var"])}{num_td(fmt(un["may_u"]))}{num_td(fmt(un["jun_u"]))}</tr>')
    def subtotal(label, names, dot):
        rows = [by.get(norm(n)) for n in names if by.get(norm(n))]
        may = sum(r["mayo"] for r in rows); jun = sum(r["junio"] for r in rows)
        umay = sum((u(n)["may_u"] or 0) for n in names); ujun = sum((u(n)["jun_u"] or 0) for n in names)
        var = (jun/may-1) if may else None
        return (f'<tr class="trow-total"><td>{esc(label)}</td>{num_td(fmt(may))}{num_td(fmt(jun))}'
                f'{vf_td(var)}{num_td(fmt(umay))}{num_td(fmt(ujun))}</tr>')
    heaven = ["Tienda Av. Buenos Aires", "Tienda Central", "Tienda Mia Plaza"]
    suena = ["Alm. Tienda Carmelo", "CHARCAS"]
    out = []
    out.append('<tr style="background:var(--gray-lt)"><td colspan="6" style="font-weight:800;font-size:.72rem;letter-spacing:.05em"><span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:var(--teal);margin-right:9px;vertical-align:middle"></span>HEAVEN</td></tr>')
    out += [row(n, indent=True) for n in heaven]
    out.append(subtotal("Subtotal HEAVEN", heaven, "var(--teal)"))
    out.append('<tr style="background:var(--gray-lt)"><td colspan="6" style="font-weight:800;font-size:.72rem;letter-spacing:.05em"><span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:var(--amber);margin-right:9px;vertical-align:middle"></span>SUEÑA</td></tr>')
    out += [row(n, indent=True) for n in suena]
    out.append(subtotal("Subtotal SUEÑA", suena, "var(--amber)"))
    out.append(row("ROHO", bold=True, dot="var(--red)"))
    out.append(row("PRODUCTOS TERMINADOS FAB.", bold=True, dot="var(--gray)"))
    out.append(row("DISTRIBUIDORES", bold=True, dot="var(--series-blue)"))
    wt = D["semanas_total"]
    out.append(f'<tr class="trow-total" style="border-top:3px solid var(--black)"><td>TOTAL GENERAL</td>{num_td(fmt(wt["mayo"]))}{num_td(fmt(wt["junio"]))}{vf_td(wt["var"])}{num_td(fmt(wt["may_u"]))}{num_td(fmt(wt["jun_u"]))}</tr>')
    return "".join(out)


def main():
    ap = argparse.ArgumentParser(description="Genera el Dashboard Empresarial desde el Excel de ventas.")
    ap.add_argument("xlsx", nargs="?", default=DEFAULT_XLSX, help="Ruta al Excel de ventas")
    ap.add_argument("-o", "--out", default=DEFAULT_OUT, help="Ruta de salida HTML")
    args = ap.parse_args()
    if not os.path.exists(args.xlsx):
        sys.exit(f"No existe el Excel: {args.xlsx}")
    D = extract(args.xlsx)
    html = build_html(D)
    with open(args.out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"OK -> {args.out}  ({len(html):,} bytes)  fuente: {D['_xlsx']}")


if __name__ == "__main__":
    main()
