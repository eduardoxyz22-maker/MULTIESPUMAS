"""
Regenera dashboard-comercial.html desde datos.xlsx (descargado de Google Drive).
Llamado por .github/workflows/actualizar-dashboard.yml
"""
import sys, re, json, base64, gzip, io, math, warnings
warnings.filterwarnings("ignore")

XLSX = "datos.xlsx"
HTML = "dashboard-comercial.html"

# ── helpers bundler ──────────────────────────────────────────────────────────

def decompress(entry):
    raw = base64.b64decode(entry["data"])
    if entry.get("compressed"):
        raw = gzip.decompress(raw)
    return raw.decode("utf-8")

def compress_encode(text):
    raw = text.encode("utf-8")
    buf = io.BytesIO()
    with gzip.GzipFile(fileobj=buf, mode="wb", mtime=0) as gz:
        gz.write(raw)
    return base64.b64encode(buf.getvalue()).decode("ascii")

def safe(v, default=0):
    if v is None:
        return default
    if isinstance(v, float) and math.isnan(v):
        return default
    if isinstance(v, str):
        try:
            return float(v)
        except ValueError:
            return default
    return v

def pct(a, b):
    a, b = safe(a), safe(b)
    return a / b if b else 0

# ── leer Excel ───────────────────────────────────────────────────────────────

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

wb = openpyxl.load_workbook(XLSX, data_only=True)
d  = wb["Dashboard"]
mg = wb["MAYO GLOBAL"]

def dc(row, col):
    return safe(d.cell(row, col).value)

def mc(row, col):
    return safe(mg.cell(row, col).value)

def mc_raw(row, col):
    return mg.cell(row, col).value

# ── período ──────────────────────────────────────────────────────────────────
# Dashboard R36 col M=13: días transcurridos  /  R37 col M=13: días totales
dias_trans   = int(dc(36, 13))
dias_totales = int(dc(37, 13))
dias_rest    = dias_totales - dias_trans

titulo = str(d.cell(1, 1).value or "")
MESES = {"ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
         "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12}
MES_NOMBRE = {v:k.capitalize() for k,v in MESES.items()}
mes_num, anio = 5, 2025
for nombre, num in MESES.items():
    if nombre in titulo.upper():
        mes_num = num
        break
for tok in titulo.split():
    if tok.isdigit() and len(tok) == 4:
        anio = int(tok)
        break
mes_nombre = MES_NOMBRE[mes_num]

# ── global KPIs ───────────────────────────────────────────────────────────────
# Dashboard R5: (ventaTotal, -, metaMensual, -, %cumpl, -, ventaTiendas, -, externos, -, comisiones)
venta_total    = dc(5, 1)
meta_mensual   = dc(5, 3)
ventas_tiendas = dc(5, 7)
externos_total = dc(5, 9)

# Dashboard R29 TOTAL comisiones: col B=comis, C=bonos, D=total
comisiones_tot = dc(29, 2)
bonos_tot      = dc(29, 3)

# Dashboard col K (11): R27=leads, R28=ventas conc, R29=conv, R30=ticket, R31=productos, R32=falta
leads_total    = int(dc(27, 11))
ventas_conc    = int(dc(28, 11))
conv_global    = dc(29, 11)
ticket_prom    = dc(30, 11)
productos_vend = int(dc(31, 11))
falta_meta     = dc(32, 11)

ritmo_diario_g  = venta_total / dias_trans if dias_trans else 0
proyeccion_g    = ritmo_diario_g * dias_totales
porc_proyeccion = pct(proyeccion_g, meta_mensual)
porc_presup     = pct(venta_total, meta_mensual)
meta_min_total  = dc(23, 3)   # TOTAL row meta mínima
porc_meta_min   = pct(venta_total, meta_min_total)

# ── tiendas ───────────────────────────────────────────────────────────────────
# Dashboard R19-22: (nombre, monto, metaMin, %min, presupuesto, %pres)
tienda_names_dash = ["SUEÑA","HEAVEN","OTROS","ROHO"]
tienda_rows_dash  = [19, 20, 21, 22]

# Diferencias vs Abril: MAYO GLOBAL R42-45 col F=6 nombre, G=7 diferencia
mom_diffs = {}
for r in range(42, 46):
    nombre_mom = str(mc_raw(r, 6) or "").strip().upper()
    diff_val   = mc(r, 7)
    if nombre_mom:
        mom_diffs[nombre_mom] = diff_val

# Comisiones por tienda: Dashboard R27=SUEÑA, R28=HEAVEN
comisiones_tiendas = {
    "SUEÑA":  {"comis": dc(27,2), "bonos": dc(27,3), "total": dc(27,4), "comisionados": int(dc(27,5))},
    "HEAVEN": {"comis": dc(28,2), "bonos": dc(28,3), "total": dc(28,4), "comisionados": int(dc(28,5))},
}

# Proyección por tienda: Dashboard R37=SUEÑA, R38=HEAVEN (col D=4)
tienda_proyecciones = {
    "SUEÑA":  dc(37, 4),
    "HEAVEN": dc(38, 4),
    "EXTERNOS": dc(39, 4),
    "OTROS": dc(39, 4),
}

tiendas_data = []
for nombre, row in zip(tienda_names_dash, tienda_rows_dash):
    monto      = dc(row, 2)
    meta_min   = dc(row, 3)
    pct_min_v  = dc(row, 4)
    presup     = dc(row, 5)
    pct_pres_v = dc(row, 6)

    diff_key      = {"SUEÑA":"SUEÑA","HEAVEN":"HEAVEN","OTROS":"EXTERNOS","ROHO":"ROHO"}.get(nombre, nombre)
    diferencia        = mom_diffs.get(diff_key, 0)
    mes_pasado_monto  = monto - diferencia if diferencia else 0
    crec_abril        = pct(diferencia, mes_pasado_monto)

    ct   = comisiones_tiendas.get(nombre, {"comis":0,"bonos":0,"total":0,"comisionados":0})
    proj = tienda_proyecciones.get(nombre)
    ritmo = monto / dias_trans if dias_trans else 0

    t_leads = None
    if nombre == "HEAVEN":
        t_leads = leads_total - (leads_total * 1320 // 2350)  # rough split
    elif nombre == "SUEÑA":
        t_leads = leads_total * 1320 // 2350

    tiendas_data.append({
        "nombre": nombre,
        "monto": monto,
        "metaMin": meta_min,
        "pctMin": pct_min_v,
        "presupuesto": presup,
        "pctPres": pct_pres_v,
        "proyeccion": proj if proj else None,
        "ritmoDiario": ritmo,
        "pctProyeccionMeta": pct(proj, meta_min) if proj and meta_min else None,
        "leads": t_leads,
        "conversion": None,
        "comisiones": ct["comis"],
        "bonos": ct["bonos"],
        "totalPagado": ct["total"],
        "comisionados": ct["comisionados"],
        "mesPasadoMonto": mes_pasado_monto if mes_pasado_monto else None,
        "crecimientoVsAbril": crec_abril if mes_pasado_monto else None,
    })

# ── vendedores ────────────────────────────────────────────────────────────────
# Dashboard R9-15: (rank, nombre, monto, pct, proyeccion, ritmo)
# MAYO GLOBAL R5-7: SUEÑA (nombre, visitas, ventas, online, productos, monto, metaMin, %min, metaPres, %pres)
# MAYO GLOBAL R16-19: HEAVEN (nombre, -, ventas, -, productos, monto, metaMin, %min, metaPres, %pres)
# MAYO GLOBAL R53-55: SUEÑA KPIs (nombre, tasaCierre, ingresoLead, comPct, crecAbril, proyeccion)
# MAYO GLOBAL R60-63: HEAVEN KPIs (nombre, conversion, ingresoLead, comPct, crecAbril, proyeccion)
# MAYO GLOBAL R70-76: Ranking (rank, nombre, monto, pctTotal, mejorMes, nuevoRecord)

sueña_detail = {}
for r in range(5, 8):
    n = str(mc_raw(r,1) or "").strip()
    if not n or "TOTAL" in n.upper():
        continue
    sueña_detail[n.upper()] = {
        "visitas":   int(safe(mc(r,2))),
        "ventas":    int(safe(mc(r,3))),
        "online":    int(safe(mc(r,4))),
        "productos": int(safe(mc(r,5))),
        "monto":     mc(r,6),
        "metaMin":   mc(r,7),
        "pctMin":    mc(r,8),
        "metaPres":  mc(r,9),
        "pctPres":   mc(r,10),
    }

heaven_detail = {}
for r in range(16, 20):
    n = str(mc_raw(r,1) or "").strip()
    if not n or "TOTAL" in n.upper():
        continue
    heaven_detail[n.upper()] = {
        "ventas_conc": int(safe(mc(r,3))),
        "productos":   int(safe(mc(r,5))),
        "monto":       mc(r,6),
        "metaMin":     mc(r,7),
        "pctMin":      mc(r,8),
        "metaPres":    mc(r,9),
        "pctPres":     mc(r,10),
    }

sueña_kpi = {}
for r in range(53, 56):
    n = str(mc_raw(r,1) or "").strip()
    if not n:
        continue
    sueña_kpi[n.upper()] = {
        "tasaCierre":       mc(r,2),
        "ingresoLead":      mc(r,3),
        "pctComision":      mc(r,4),
        "crecimientoAbril": mc(r,5),
        "proyeccion":       mc(r,6),
    }

heaven_kpi = {}
for r in range(60, 64):
    n = str(mc_raw(r,1) or "").strip()
    if not n:
        continue
    heaven_kpi[n.upper()] = {
        "conversion":       mc(r,2),
        "ingresoLead":      mc(r,3),
        "pctComision":      mc(r,4),
        "crecimientoAbril": mc(r,5),
        "proyeccion":       mc(r,6),
    }

ranking = {}
for r in range(70, 77):
    n = str(mc_raw(r,2) or "").strip()
    if not n:
        continue
    rec_raw = mc_raw(r,6)
    ranking[n.upper()] = {
        "mejorMes":   int(safe(mc(r,5))),
        "nuevoRecord": bool(rec_raw and str(rec_raw).strip()),
    }

# Build vendedores list
SUEÑA_NAMES  = {"JUAN PABLO","FERNANDO","MAURICIO"}
HEAVEN_NAMES = {"MARIA","ISABEL","CAROLA","MIRIAN"}
TIENDA_MAP   = {**{n:"SUEÑA" for n in SUEÑA_NAMES}, **{n:"HEAVEN" for n in HEAVEN_NAMES}}

sueña_leads_per = 440   # ~1320/3 (constant unless sheet changes)

vendedores_data = []
for r in range(9, 16):
    rank = d.cell(r,1).value
    if not isinstance(rank, (int, float)):
        continue
    nombre    = str(d.cell(r,2).value or "").strip()
    nom_up    = nombre.upper()
    tienda    = TIENDA_MAP.get(nom_up, "HEAVEN")
    monto     = safe(d.cell(r,3).value)
    pct_total = safe(d.cell(r,4).value)

    if tienda == "SUEÑA":
        det = sueña_detail.get(nom_up, {})
        kpi = sueña_kpi.get(nom_up, {})
        rnk = ranking.get(nom_up, {})
        leads        = sueña_leads_per
        visitas      = det.get("visitas")
        ventas       = det.get("ventas")
        online       = det.get("online")
        productos    = det.get("productos")
        meta_min     = det.get("metaMin", 0)
        pct_min_v    = det.get("pctMin", 0)
        meta_pres    = det.get("metaPres", 0)
        pct_pres_v   = det.get("pctPres", 0)
        pct_comision = kpi.get("pctComision", 0)
        ingreso_lead = kpi.get("ingresoLead", 0)
        crec_abril   = kpi.get("crecimientoAbril", 0)
        comision     = round(monto * pct_comision, 4) if pct_comision else 0
        ticket_v     = monto / ventas if ventas else 0
        prod_venta   = productos / ventas if (productos and ventas) else 0
        venta_dia    = monto / dias_trans if dias_trans else 0
        vend = {
            "nombre": nombre, "tienda": tienda,
            "monto": monto, "pctTotal": pct_total,
            "proyeccion": kpi.get("proyeccion") or safe(d.cell(r,5).value),
            "ritmoDiario": safe(d.cell(r,6).value),
            "visitas": visitas, "ventas": ventas, "online": online, "productos": productos,
            "leads": leads, "conversion": kpi.get("tasaCierre", 0), "conversionVisita": kpi.get("tasaCierre"),
            "metaMin": meta_min, "pctMin": pct_min_v, "metaPres": meta_pres, "pctPres": pct_pres_v,
            "ticketProm": ticket_v, "prodPorVenta": prod_venta, "ventasPorDia": venta_dia,
            "ingresoLead": ingreso_lead, "comision": comision, "pctComision": pct_comision,
            "bonoTitanio": 0,
            "crecimientoVsAbril": crec_abril,
            "mejorMes": rnk.get("mejorMes", 0), "nuevoRecord": rnk.get("nuevoRecord", False),
            "tasaCierre": kpi.get("tasaCierre"),
        }
    else:  # HEAVEN
        det = heaven_detail.get(nom_up, {})
        kpi = heaven_kpi.get(nom_up, {})
        rnk = ranking.get(nom_up, {})
        ingreso_lead = kpi.get("ingresoLead", 0)
        leads        = round(monto / ingreso_lead) if ingreso_lead else 265
        visitas      = det.get("ventas_conc")
        online       = det.get("productos")
        meta_min     = det.get("metaMin", 0)
        pct_min_v    = det.get("pctMin", 0)
        meta_pres    = det.get("metaPres", 0)
        pct_pres_v   = det.get("pctPres", 0)
        pct_comision = kpi.get("pctComision", 0)
        crec_abril   = kpi.get("crecimientoAbril", 0)
        ventas_c     = det.get("ventas_conc", 0)
        productos_v  = det.get("productos", 0)
        comision     = round(monto * pct_comision, 4) if pct_comision else 0
        ticket_v     = monto / ventas_c if ventas_c else 0
        prod_venta   = productos_v / ventas_c if ventas_c else 0
        venta_dia    = monto / dias_trans if dias_trans else 0
        vend = {
            "nombre": nombre, "tienda": tienda,
            "monto": monto, "pctTotal": pct_total,
            "proyeccion": kpi.get("proyeccion") or safe(d.cell(r,5).value),
            "ritmoDiario": safe(d.cell(r,6).value),
            "visitas": visitas, "ventas": None, "online": online, "productos": None,
            "leads": leads, "conversion": kpi.get("conversion", 0), "conversionVisita": None,
            "metaMin": meta_min, "pctMin": pct_min_v, "metaPres": meta_pres, "pctPres": pct_pres_v,
            "ticketProm": ticket_v, "prodPorVenta": prod_venta, "ventasPorDia": venta_dia,
            "ingresoLead": ingreso_lead, "comision": comision, "pctComision": pct_comision,
            "bonoTitanio": 0,
            "crecimientoVsAbril": crec_abril,
            "mejorMes": rnk.get("mejorMes", 0), "nuevoRecord": rnk.get("nuevoRecord", False),
        }
    vendedores_data.append(vend)

# ── clientes externos ─────────────────────────────────────────────────────────
# MAYO GLOBAL R29-36: (nombre, -, productos, monto)
clientes_raw = []
total_ext = externos_total or 1
for r in range(29, 37):
    nombre_c = str(mc_raw(r, 1) or "").strip()
    if not nombre_c or nombre_c.upper() in ("TOTAL","VENDEDOR"):
        continue
    monto_c  = mc(r, 4)
    ventas_c = int(safe(mc(r, 3)))
    clientes_raw.append({
        "nombre": nombre_c,
        "ventas": ventas_c,
        "monto":  monto_c,
        "pct":    pct(monto_c, total_ext),
    })

# ── armar __DATA__ ────────────────────────────────────────────────────────────

data_js = f"""// Datos extraídos de "{XLSX}" — hoja Dashboard + MAYO GLOBAL
// {mes_nombre} {anio} — Día {dias_trans} de {dias_totales}

window.__DATA__ = {{
  periodo: {{ mes: "{mes_nombre}", anio: {anio}, diasTranscurridos: {dias_trans}, diasTotales: {dias_totales}, diasRestantes: {dias_rest} }},

  global: {{
    ventaTotal: {venta_total},
    presupuesto: {meta_mensual},
    porcPresupuesto: {porc_presup},
    metaMinTotal: {meta_min_total},
    porcMetaMin: {porc_meta_min},
    leadsTotal: {leads_total},
    ventasConcretadas: {ventas_conc},
    conversionGlobal: {conv_global},
    ticketPromGlobal: {ticket_prom},
    productosVendidos: {productos_vend},
    faltaParaMeta: {falta_meta},
    comisionesTotales: {comisiones_tot},
    bonosTotales: {bonos_tot},
    ritmoDiario: {ritmo_diario_g},
    proyeccionCierre: {proyeccion_g},
    porcProyeccion: {porc_proyeccion},
  }},

  vendedores: {json.dumps(vendedores_data, ensure_ascii=False, indent=4)},

  tiendas: {json.dumps(tiendas_data, ensure_ascii=False, indent=4)},

  clientes: {json.dumps(clientes_raw, ensure_ascii=False, indent=4)},
}};"""

# ── actualizar dashboard-comercial.html ──────────────────────────────────────

content = open(HTML, encoding="utf-8").read()
manifest_m = re.search(r'<script type="__bundler/manifest">(.*?)</script>', content, re.DOTALL)
if not manifest_m:
    print("ERROR: no se encontró el manifest en el HTML")
    sys.exit(1)

manifest = json.loads(manifest_m.group(1).strip())

data_uuid = next((k for k in manifest if k.startswith("eb1309ec")), None)
if not data_uuid:
    print("ERROR: no se encontró el asset eb1309ec en el manifest")
    sys.exit(1)

manifest[data_uuid]["data"] = compress_encode(data_js)
manifest[data_uuid]["compressed"] = True

new_manifest_json = json.dumps(manifest, separators=(",", ":"))
new_content = (content[:manifest_m.start(1)] + "\n" +
               new_manifest_json + "\n  " + content[manifest_m.end(1):])

with open(HTML, "w", encoding="utf-8") as f:
    f.write(new_content)

print(f"OK — Dashboard actualizado: {mes_nombre} {anio}, día {dias_trans}/{dias_totales}")
print(f"     Venta total: {venta_total:,.0f} / Meta: {meta_mensual:,.0f} ({porc_presup*100:.1f}%)")
print(f"     Vendedores: {len(vendedores_data)} | Tiendas: {len(tiendas_data)} | Clientes: {len(clientes_raw)}")
